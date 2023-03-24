import os
from openpyxl import Workbook, load_workbook
import pandas as pd


def saveNoRe(inp):
    #inp = input("You: ")

    noRe = "D:\Code\Project\\noRp\\noResponses.xlsx"

    if not os.path.exists(noRe):
        # Tạo file mới
        workbook = Workbook()
        worksheet = workbook.active
        # Đặt tên cho các cột
        worksheet.cell(row=1, column=1).value = "Tag"
        worksheet.cell(row=1, column=2).value = "Patterns"
        worksheet.cell(row=1, column=3).value = "Responses"
        workbook.save(noRe)
    else:
        # Mở file có sẵn
        workbook = load_workbook(filename=noRe)
        worksheet = workbook.active

    row_num = worksheet.max_row+1
    # Thêm câu mới vào file
    worksheet.cell(row=row_num, column=2).value = inp

    row_num += 1
    # Lưu file workbook
    workbook.save(noRe)

